# -*- coding: utf-8 -*-
"""
CeloDocs extractor + email sender (attachments)
- Genera JSON + XLSX como ya lo haces
- Opcional: empaqueta out_celodocs/sql_full/*.sql en ZIP
- Envía correo vía SMTP (Gmail/Relay) con adjuntos

Credenciales sugeridas por env vars:
  EMAIL_USER, EMAIL_PASS
  SMTP_SERVER (default smtp.gmail.com)
  SMTP_PORT   (default 587)
Opciones email:
  CELODOCS_EMAIL_TO        (requerido para enviar)
  CELODOCS_EMAIL_CC        (opcional, separado por comas)
  CELODOCS_EMAIL_BCC       (opcional, separado por comas)
  CELODOCS_ATTACH_SQL_ZIP  ("1" para adjuntar zip de sql_full)
"""

import os
import json
import datetime
import configparser
from pathlib import Path
from typing import Any, Dict, List, Optional
import glob
import re
import zipfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from pycelonis import get_celonis

# Usa tu BusinessGraphAPI real
from app.src.ocpm.business_graph_api import BusinessGraphAPI
from dotenv import load_dotenv

# Carga .env (por defecto busca en el directorio actual)
load_dotenv()

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_GRAY  = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
FONT_BOLD  = Font(bold=True)

# =========================
# CONFIG
# =========================
TEMPLATE_XLSX = "Requirements Template (OCPM).xlsx"

OUT_DIR = Path("out_celodocs")
OUT_DIR.mkdir(parents=True, exist_ok=True)

SQL_DIR = OUT_DIR / "sql_full"
SQL_DIR.mkdir(parents=True, exist_ok=True)

EXCEL_SQL_PREVIEW_CHARS = 32000  # preview seguro para Excel (límite ~32767)

DEFAULT_NAMESPACES = ["custom", "catalog", "standard", "sap", "celonis"]

SHEETS_HEADERS = {
    "Objects": [
        "ObjectTypeName",
        "ObjectTypeId",
        "Namespace",
        "DisplayName",
        "Description",
        "SourceTable",
        "PrimaryKey",
        "SQL_FactoryId",
        "SQL_Statement",
    ],
    "Events": [
        "EventTypeName",
        "EventTypeId",
        "Namespace",
        "DisplayName",
        "Description",
        "SourceTable",
        "ActivityColumn",
        "TimestampColumn",
        "RelatedObjects",
        "SQL_FactoryId",
        "SQL_Statement",
    ],
    "SQL_Factories": [
        "EntityName",
        "EntityKind",
        "FactoryId",
        "DataConnectionId",
        "Namespace",
        "DisplayName",
        "Draft",
        "TargetKind",
        "SQL_Statement",
    ],
    "Relationships": [
        "FromObject",
        "ToObject",
        "Cardinality",
        "JoinKeys",
        "Description",
    ],
}

# =========================
# EMAIL helpers
# =========================
def _split_emails(s: Optional[str]) -> List[str]:
    if not s:
        return []
    return [x.strip() for x in str(s).split(",") if x.strip()]

def send_email_with_attachments(
    to_emails: List[str],
    subject: str,
    body_text: str,
    attachments: List[Path],
    cc_emails: Optional[List[str]] = None,
    bcc_emails: Optional[List[str]] = None,
    smtp_server: Optional[str] = None,
    smtp_port: Optional[int] = None,
    sender_email: Optional[str] = None,
    sender_password: Optional[str] = None,
) -> tuple[bool, str]:
    """
    Envía correo con adjuntos.
    - Usa STARTTLS.
    - Para Gmail: usa App Password (recomendado) o SMTP corporativo.
    """
    cc_emails = cc_emails or []
    bcc_emails = bcc_emails or []

    sender_email = sender_email or os.getenv("EMAIL_USER")
    sender_password = sender_password or os.getenv("EMAIL_PASS")

    smtp_server = smtp_server or os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(smtp_port or os.getenv("SMTP_PORT", "587"))

    if not sender_email or not sender_password:
        return False, "Missing EMAIL_USER / EMAIL_PASS env vars (or passed args)."

    if not to_emails:
        return False, "No recipients (to_emails) provided."

    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = ", ".join(to_emails)
    if cc_emails:
        msg["Cc"] = ", ".join(cc_emails)
    msg["Subject"] = subject

    msg.attach(MIMEText(body_text, "plain", "utf-8"))

    # Adjuntos
    for p in attachments:
        if not p:
            continue
        p = Path(p)
        if not p.exists():
            return False, f"Attachment not found: {p}"

        part = MIMEBase("application", "octet-stream")
        with open(p, "rb") as f:
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{p.name}"')
        msg.attach(part)

    # Enviar
    all_rcpt = list(dict.fromkeys(to_emails + cc_emails + bcc_emails))  # unique, keep order
    try:
        with smtplib.SMTP(smtp_server, smtp_port, timeout=60) as server:
            server.ehlo()
            server.starttls()
            server.ehlo()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, all_rcpt, msg.as_string())
        return True, f"Sent to: {', '.join(all_rcpt)}"
    except Exception as e:
        return False, f"SMTP send failed: {e}"

def zip_sql_dir(sql_dir: Path, out_zip: Path) -> Path:
    """
    Empaqueta todos los .sql dentro de sql_dir (recursivo) en un zip.
    """
    sql_dir = Path(sql_dir)
    out_zip = Path(out_zip)
    if not sql_dir.exists():
        raise FileNotFoundError(f"SQL dir not found: {sql_dir}")

    with zipfile.ZipFile(out_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in sql_dir.rglob("*.sql"):
            arcname = p.relative_to(sql_dir.parent)  # incluye sql_full/...
            zf.write(p, arcname.as_posix())
    return out_zip

# =========================
# Excel helpers
# =========================
def safe_file_slug(s: str, max_len: int = 80) -> str:
    s = re.sub(r"[^\w\-]+", "_", str(s).strip(), flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")
    return (s[:max_len] or "entity")

def write_full_sql_file(entity_kind: str, entity_name: str, factory_id: str, sql_text: str) -> Path:
    kind = safe_file_slug(entity_kind.upper(), 20)
    name = safe_file_slug(entity_name, 80)
    fid = safe_file_slug(factory_id, 60) if factory_id else "no_factory"
    path = SQL_DIR / f"{kind}__{name}__{fid}.sql"
    path.write_text(sql_text or "", encoding="utf-8")
    return path

def excel_hyperlink(path: Path, label: str = "Open SQL") -> str:
    p = str(path.resolve()).replace("\\", "/")
    return f'=HYPERLINK("{p}","{label}")'

def normalize_sql_for_diff(sql: Any) -> Any:
    if sql is None:
        return None
    s = str(sql)
    s = s.replace("\r\n", "\n").replace("\r", "\n").strip()
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s

def ensure_sheet(wb, name: str, headers: List[str]):
    if name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(name)
    if ws.max_row == 1 and ws["A1"].value is None:
        for c, h in enumerate(headers, start=1):
            ws.cell(row=1, column=c, value=h)
    return ws

def append_rows(ws, headers: List[str], rows: List[Dict[str, Any]]):
    header_to_col = {}
    max_col = ws.max_column
    for c in range(1, max_col + 1):
        v = ws.cell(row=1, column=c).value
        if v:
            header_to_col[str(v).strip()] = c
    for h in headers:
        if h not in header_to_col:
            max_col += 1
            ws.cell(row=1, column=max_col, value=h)
            header_to_col[h] = max_col

    start_row = ws.max_row + 1
    for i, row in enumerate(rows):
        r = start_row + i
        for h in headers:
            ws.cell(row=r, column=header_to_col[h], value=row.get(h))

# =========================
# Normalización entities
# =========================
def _get(obj, attr: str, default=None):
    if hasattr(obj, attr):
        return getattr(obj, attr)
    if isinstance(obj, dict):
        return obj.get(attr, default)
    return default

def object_to_row(o, factory_id, sql_clean_preview, sql_raw, sql_full_path=None, sql_full_link=None):
    return {
        "ObjectTypeName": _get(o, "name"),
        "ObjectTypeId": _get(o, "id"),
        "Namespace": _get(o, "namespace"),
        "DisplayName": _get(o, "display_name") or _get(o, "displayName"),
        "Description": _get(o, "description"),
        "SourceTable": _get(o, "source_table") or _get(o, "sourceTable"),
        "PrimaryKey": _get(o, "primary_key") or _get(o, "primaryKey"),
        "SQL_FactoryId": factory_id,
        "SQL_Statement": sql_raw,
        "SQL_Transformation_Clean": sql_clean_preview,
        "SQL_Full_Path": str(sql_full_path) if sql_full_path else None,
        "SQL_Full_Link": sql_full_link,
    }

def event_to_row(e, factory_id, sql_clean_preview, sql_raw, sql_full_path=None, sql_full_link=None):
    rel = _get(e, "related_object_types") or _get(e, "relatedObjectTypes")
    if isinstance(rel, list):
        rel = ",".join([str(x) for x in rel])
    return {
        "EventTypeName": _get(e, "name"),
        "EventTypeId": _get(e, "id"),
        "Namespace": _get(e, "namespace"),
        "DisplayName": _get(e, "display_name") or _get(e, "displayName"),
        "Description": _get(e, "description"),
        "SourceTable": _get(e, "source_table") or _get(e, "sourceTable"),
        "ActivityColumn": _get(e, "activity_column") or _get(e, "activityColumn"),
        "TimestampColumn": _get(e, "timestamp_column") or _get(e, "timestampColumn"),
        "RelatedObjects": rel,
        "SQL_FactoryId": factory_id,
        "SQL_Statement": sql_raw,
        "SQL_Transformation_Clean": sql_clean_preview,
        "SQL_Full_Path": str(sql_full_path) if sql_full_path else None,
        "SQL_Full_Link": sql_full_link,
    }

def factory_to_row(entity_name: str, entity_kind: str, f, sql: Optional[str]) -> Dict[str, Any]:
    target = _get(f, "target")
    target_kind = _get(target, "kind") if target else None
    return {
        "EntityName": entity_name,
        "EntityKind": entity_kind,
        "FactoryId": _get(f, "factory_id") or _get(f, "factoryId") or _get(f, "id"),
        "DataConnectionId": _get(f, "data_connection_id") or _get(f, "dataConnectionId"),
        "Namespace": _get(f, "namespace"),
        "DisplayName": _get(f, "display_name") or _get(f, "displayName"),
        "Draft": _get(f, "draft"),
        "TargetKind": target_kind,
        "SQL_Statement": sql,
    }

def extract_clean_sql_from_sql_factory(sql_factory_full) -> dict:
    out = {"main_sql": None, "relationship_sql": {}}
    txs = getattr(sql_factory_full, "transformations", None)
    if not txs:
        return out
    t0 = txs[0]

    def collect_sql(datasets) -> list[str]:
        parts = []
        for ds in datasets or []:
            s = getattr(ds, "sql", None)
            if s and isinstance(s, str):
                s = s.strip()
                if s:
                    parts.append(s)
        return parts

    change_parts = collect_sql(getattr(t0, "change_sql_factory_datasets", None))
    prop_parts = collect_sql(getattr(t0, "property_sql_factory_datasets", None))
    main_parts = change_parts + prop_parts
    out["main_sql"] = "\n\n-- =====================\n\n".join(main_parts) if main_parts else None

    rel_txs = getattr(t0, "relationship_transformations", None) or []
    for rel in rel_txs:
        rel_name = getattr(rel, "relationship_name", None) or "RELATIONSHIP"
        rel_parts = collect_sql(getattr(rel, "sql_factory_datasets", None))
        if rel_parts:
            out["relationship_sql"][rel_name] = "\n\n-- =====================\n\n".join(rel_parts)
    return out

def shrink_sql(sql: str, max_chars: int = 30000) -> str:
    if not sql:
        return sql
    sql = sql.strip()
    if len(sql) <= max_chars:
        return sql
    return sql[:max_chars] + "\n\n-- [TRUNCATED]"

def find_previous_report(out_dir: str, pattern: str = "Requirements_Filled_*.xlsx") -> str | None:
    files = sorted(glob.glob(os.path.join(out_dir, pattern)))
    return files[-1] if files else None

def read_sheet_as_dict(xlsx_path: str, sheet_name: str, key_col: str) -> dict[str, dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    try:
        key_idx = headers.index(key_col) + 1
    except ValueError:
        return {}

    data = {}
    for r in range(2, ws.max_row + 1):
        key = ws.cell(row=r, column=key_idx).value
        if key is None or str(key).strip() == "":
            continue
        row_dict = {}
        for c, h in enumerate(headers, start=1):
            if not h:
                continue
            row_dict[h] = ws.cell(row=r, column=c).value
        data[str(key)] = row_dict
    return data

def diff_records(before: dict[str, dict], after: dict[str, dict], compare_cols: list[str]) -> dict:
    before_keys = set(before.keys())
    after_keys = set(after.keys())
    added = sorted(list(after_keys - before_keys))
    removed = sorted(list(before_keys - after_keys))

    modified = []
    common = before_keys & after_keys
    for k in sorted(common):
        changes = {}
        b = before[k]
        a = after[k]
        for col in compare_cols:
            bv = b.get(col)
            av = a.get(col)
            if isinstance(bv, str): bv = bv.strip()
            if isinstance(av, str): av = av.strip()
            if col in ("SQL_Transformation_Clean", "SQL_Statement"):
                bv = normalize_sql_for_diff(bv)
                av = normalize_sql_for_diff(av)
            if bv != av:
                changes[col] = {"before": bv, "after": av}
        if changes:
            modified.append({"key": k, "changes": changes})
    return {"added": added, "removed": removed, "modified": modified}

def ensure_sheet_first(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
        wb.remove(ws)
    ws = wb.create_sheet(name, 0)
    return ws

def write_exec_summary(
    wb,
    prev_path: str,
    curr_path: str,
    obj_diff: dict,
    evt_diff: dict,
    before_obj: dict[str, dict],
    after_obj: dict[str, dict],
    before_evt: dict[str, dict],
    after_evt: dict[str, dict],
):
    ws = ensure_sheet_first(wb, "Executive_Summary")

    # Header
    ws["A1"] = "CeloDocs | Executive Summary (Diff)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Previous file:"
    ws["B2"] = prev_path
    ws["A3"] = "Current file:"
    ws["B3"] = curr_path
    ws["A5"] = "Summary"
    ws["A5"].font = FONT_BOLD

    # Metrics
    ws["A6"] = "Objects - Added";    ws["B6"] = len(obj_diff["added"])
    ws["A7"] = "Objects - Removed";  ws["B7"] = len(obj_diff["removed"])
    ws["A8"] = "Objects - Modified"; ws["B8"] = len(obj_diff["modified"])

    ws["D6"] = "Events - Added";     ws["E6"] = len(evt_diff["added"])
    ws["D7"] = "Events - Removed";   ws["E7"] = len(evt_diff["removed"])
    ws["D8"] = "Events - Modified";  ws["E8"] = len(evt_diff["modified"])

    # Table header for detailed changes
    start = 10
    headers = ["EntityType", "ID", "Name", "Field", "Before", "After"]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=i, value=h)
        cell.font = FONT_BOLD
        cell.fill = FILL_GRAY

    row = start + 1

    def _get_name(entity_type: str, key: str) -> str:
        # Prefer "after" name; fallback to "before"
        if entity_type == "OBJECT":
            a = after_obj.get(key) or {}
            b = before_obj.get(key) or {}
            return str(a.get("ObjectTypeName") or b.get("ObjectTypeName") or "")
        else:
            a = after_evt.get(key) or {}
            b = before_evt.get(key) or {}
            return str(a.get("EventTypeName") or b.get("EventTypeName") or "")

    def add_change_rows(entity_type: str, diff: dict):
        nonlocal row

        # Added
        for k in diff["added"]:
            ws.cell(row=row, column=1, value=entity_type)
            ws.cell(row=row, column=2, value=k)
            ws.cell(row=row, column=3, value=_get_name(entity_type, k))
            ws.cell(row=row, column=4, value="(NEW)")
            c_after = ws.cell(row=row, column=6, value="Added")
            c_after.fill = FILL_GREEN
            row += 1

        # Removed
        for k in diff["removed"]:
            ws.cell(row=row, column=1, value=entity_type)
            ws.cell(row=row, column=2, value=k)
            ws.cell(row=row, column=3, value=_get_name(entity_type, k))
            ws.cell(row=row, column=4, value="(REMOVED)")
            c_before = ws.cell(row=row, column=5, value="Removed")
            c_before.fill = FILL_RED
            row += 1

        # Modified
        for item in diff["modified"]:
            k = item["key"]
            name = _get_name(entity_type, k)
            for field, ch in item["changes"].items():
                ws.cell(row=row, column=1, value=entity_type)
                ws.cell(row=row, column=2, value=k)
                ws.cell(row=row, column=3, value=name)
                ws.cell(row=row, column=4, value=field)

                c_before = ws.cell(row=row, column=5, value=str(ch["before"]))
                c_after  = ws.cell(row=row, column=6, value=str(ch["after"]))
                c_before.fill = FILL_RED
                c_after.fill  = FILL_GREEN
                row += 1

    add_change_rows("OBJECT", obj_diff)
    add_change_rows("EVENT", evt_diff)

    # Auto width simple
    for col in range(1, 7):
        max_len = 0
        for r in range(1, row):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(12, max_len + 2))

def _ensure_headers_once(headers: List[str], extras: List[str]) -> List[str]:
    # evita duplicar columnas si el script corre varias veces en un mismo intérprete
    out = list(headers)
    for x in extras:
        if x not in out:
            out.append(x)
    return out

# =========================
# MAIN
# =========================
def main():
    # 1) Leer config.cfg
    cfg_path = os.getenv("CELODOCS_CFG_PATH", "config.cfg")
    if not os.path.exists(cfg_path):
        raise FileNotFoundError(f"No se encontró {cfg_path}")

    config = configparser.ConfigParser()
    config.read(cfg_path)

    base_url = config.get("INPUT", "url").strip()
    token = config.get("INPUT", "token").strip()
    key_type = config.get("INPUT", "key_type").strip()
    data_pool_ids_raw = config.get("INPUT", "data_pool_ids_to_extract_from", fallback="[]")
    try:
        data_pool_ids = json.loads(data_pool_ids_raw)
    except Exception:
        data_pool_ids = []

    environment = os.getenv("CELODOCS_ENVIRONMENT", "develop")
    namespaces = os.getenv("CELODOCS_NAMESPACES", "")
    namespaces_list = [x.strip() for x in namespaces.split(",") if x.strip()] or DEFAULT_NAMESPACES

    # 2) Conectar a Celonis
    celonis = get_celonis(base_url=base_url, api_token=token, key_type=key_type)

    # 3) workspace_id = primer data pool
    if not data_pool_ids:
        dp0 = celonis.data_integration.get_data_pools()[0]
    else:
        dp0 = celonis.data_integration.get_data_pool(data_pool_ids[0])
    workspace_id = dp0.id

    api = BusinessGraphAPI(celonis.client, environment=environment, workspace_id=workspace_id)

    # 4) Data sources (si los necesitas después)
    _ = api.get_factories_data_sources()

    # 5) Objetos y Eventos
    objects = api.get_object_entities({}, namespace=namespaces_list)
    events  = api.get_event_entities({}, namespace=namespaces_list)

    # 6) Factories
    factories_by_entity: Dict[str, List[Any]] = {}
    api.get_all_factories_entities(factories_by_entity)

    def pick_factory(f_list: List[Any]) -> Optional[Any]:
        if not f_list:
            return None
        non_draft = [f for f in f_list if not bool(getattr(f, "draft", False))]
        return non_draft[0] if non_draft else f_list[0]

    objects_rows: List[Dict[str, Any]] = []
    events_rows:  List[Dict[str, Any]] = []
    factories_rows: List[Dict[str, Any]] = []

    # ---- OBJECTS
    for name, o in objects.items():
        f_list = factories_by_entity.get(name, [])
        f = pick_factory(f_list)

        factory_id = None
        sql_raw = None
        sql_full = None
        sql_preview = None
        sql_path = None
        sql_link = None

        if f:
            factory_id = getattr(f, "factory_id", None) or getattr(f, "factoryId", None) or getattr(f, "id", None)
            try:
                sql_factory_full = api.get_sql_transform_from_factory(factory_id)
                sql_raw = getattr(sql_factory_full, "sql", None) or getattr(sql_factory_full, "statement", None)

                clean = extract_clean_sql_from_sql_factory(sql_factory_full)
                sql_full = clean["main_sql"] or sql_raw or ""

                sql_path = write_full_sql_file("OBJECT", name, str(factory_id), sql_full)
                sql_link = excel_hyperlink(sql_path, "Open SQL")
                sql_preview = shrink_sql(sql_full, max_chars=EXCEL_SQL_PREVIEW_CHARS)
            except Exception as ex:
                sql_full = f"-- ERROR getting SQL for {factory_id}: {ex}"
                sql_path = write_full_sql_file("OBJECT", name, str(factory_id), sql_full)
                sql_link = excel_hyperlink(sql_path, "Open SQL")
                sql_preview = shrink_sql(sql_full, max_chars=EXCEL_SQL_PREVIEW_CHARS)

            factories_rows.append(factory_to_row(name, "OBJECT", f, sql_preview))

        objects_rows.append(object_to_row(o, factory_id, sql_preview, sql_raw, sql_path, sql_link))

    # ---- EVENTS
    for name, e in events.items():
        f_list = factories_by_entity.get(name, [])
        f = pick_factory(f_list)

        factory_id = None
        sql_raw = None
        sql_full = None
        sql_preview = None
        sql_path = None
        sql_link = None

        if f:
            factory_id = getattr(f, "factory_id", None) or getattr(f, "factoryId", None) or getattr(f, "id", None)
            try:
                sql_factory_full = api.get_sql_transform_from_factory(factory_id)
                sql_raw = getattr(sql_factory_full, "sql", None) or getattr(sql_factory_full, "statement", None)

                clean = extract_clean_sql_from_sql_factory(sql_factory_full)
                sql_full = clean["main_sql"] or sql_raw or ""

                sql_path = write_full_sql_file("EVENT", name, str(factory_id), sql_full)
                sql_link = excel_hyperlink(sql_path, "Open SQL")
                sql_preview = shrink_sql(sql_full, max_chars=EXCEL_SQL_PREVIEW_CHARS)
            except Exception as ex:
                sql_full = f"-- ERROR getting SQL for {factory_id}: {ex}"
                sql_path = write_full_sql_file("EVENT", name, str(factory_id), sql_full)
                sql_link = excel_hyperlink(sql_path, "Open SQL")
                sql_preview = shrink_sql(sql_full, max_chars=EXCEL_SQL_PREVIEW_CHARS)

            factories_rows.append(factory_to_row(name, "EVENT", f, sql_preview))

        events_rows.append(event_to_row(e, factory_id, sql_preview, sql_raw, sql_path, sql_link))

    # 8) Guardar JSON snapshot
    tag = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    json_out = OUT_DIR / f"ocpm_requirements_snapshot_{tag}.json"
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(
            {
                "environment": environment,
                "workspace_id": workspace_id,
                "namespaces": namespaces_list,
                "objects_count": len(objects_rows),
                "events_count": len(events_rows),
                "objects": objects_rows,
                "events": events_rows,
                "sql_factories": factories_rows,
            },
            f,
            ensure_ascii=False,
            indent=2,
            default=str,
        )

    # 9) Exportar a Excel template
    from openpyxl import Workbook

    template_path = Path(TEMPLATE_XLSX)
    if not template_path.exists():
        raise FileNotFoundError(f"No encontré el template: {TEMPLATE_XLSX}")

    try:
        wb = openpyxl.load_workbook(template_path)
    except Exception as e:
        print(f"[WARN] No pude abrir el template (posible corrupción): {e}")
        print("[WARN] Generando un workbook nuevo con las hojas estándar...")
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # ✅ evita duplicar headers por mutación global
    obj_headers = _ensure_headers_once(SHEETS_HEADERS["Objects"], ["SQL_Transformation_Clean", "SQL_Full_Path", "SQL_Full_Link"])
    evt_headers = _ensure_headers_once(SHEETS_HEADERS["Events"],  ["SQL_Transformation_Clean", "SQL_Full_Path", "SQL_Full_Link"])

    ws_obj = ensure_sheet(wb, "Objects", obj_headers)
    ws_evt = ensure_sheet(wb, "Events", evt_headers)
    ws_fac = ensure_sheet(wb, "SQL_Factories", SHEETS_HEADERS["SQL_Factories"])
    ws_rel = ensure_sheet(wb, "Relationships", SHEETS_HEADERS["Relationships"])

    append_rows(ws_obj, obj_headers, objects_rows)
    append_rows(ws_evt, evt_headers, events_rows)
    append_rows(ws_fac, SHEETS_HEADERS["SQL_Factories"], factories_rows)
    _ = ws_rel  # reservado

    out_xlsx = OUT_DIR / f"Requirements_Filled_{tag}.xlsx"

    # --- Diff contra versión anterior ---
    prev = find_previous_report(str(OUT_DIR), pattern="Requirements_Filled_*.xlsx")
    if prev:
        before_obj = read_sheet_as_dict(prev, "Objects", key_col="ObjectTypeId")
        before_evt = read_sheet_as_dict(prev, "Events", key_col="EventTypeId")

        after_obj = {str(r["ObjectTypeId"]): r for r in objects_rows if r.get("ObjectTypeId")}
        after_evt = {str(r["EventTypeId"]): r for r in events_rows  if r.get("EventTypeId")}

        obj_cols = ["ObjectTypeName", "Namespace", "DisplayName", "SourceTable", "PrimaryKey", "SQL_Transformation_Clean"]
        evt_cols = ["EventTypeName", "Namespace", "DisplayName", "SourceTable", "ActivityColumn", "TimestampColumn", "SQL_Transformation_Clean"]

        obj_diff = diff_records(before_obj, after_obj, obj_cols)
        evt_diff = diff_records(before_evt, after_evt, evt_cols)

        write_exec_summary(
            wb,
            prev_path=os.path.basename(prev),
            curr_path=os.path.basename(str(out_xlsx)),
            obj_diff=obj_diff,
            evt_diff=evt_diff,
            before_obj=before_obj,
            after_obj=after_obj,
            before_evt=before_evt,
            after_evt=after_evt,
        )
    else:
        ws = ensure_sheet_first(wb, "Executive_Summary")
        ws["A1"] = "CeloDocs | Executive Summary (Diff)"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A3"] = "No previous report found. This is the first generated version."

    wb.save(out_xlsx)

    # =========================
    # EMAIL + ADJUNTOS
    # =========================
    subject = f"CeloDocs | Requirements OCPM (Objects+Events) | env={environment} | {tag}"

    to_emails  = _split_emails(os.getenv("CELODOCS_EMAIL_TO"))
    cc_emails  = _split_emails(os.getenv("CELODOCS_EMAIL_CC"))
    bcc_emails = _split_emails(os.getenv("CELODOCS_EMAIL_BCC"))

    attach_sql_zip = os.getenv("CELODOCS_ATTACH_SQL_ZIP", "0").strip() in ("1", "true", "TRUE", "yes", "YES")

    attachments = [out_xlsx, json_out]

    sql_zip_path = None
    if attach_sql_zip:
        sql_zip_path = OUT_DIR / f"sql_full_{tag}.zip"
        zip_sql_dir(SQL_DIR, sql_zip_path)
        attachments.append(sql_zip_path)

    body = "\n".join(
        [
            "Hola,",
            "",
            "Adjunto el reporte generado por CeloDocs.",
            f"- Environment: {environment}",
            f"- WorkspaceId: {workspace_id}",
            f"- Namespaces: {', '.join(namespaces_list)}",
            f"- Objects: {len(objects_rows)}",
            f"- Events: {len(events_rows)}",
            "",
            f"Archivos adjuntos:",
            f"- {out_xlsx.name}",
            f"- {json_out.name}",
            f"- {sql_zip_path.name}" if sql_zip_path else "- (sql_full zip no adjuntado)",
            "",
            "Saludos,",
        ]
    )

    if to_emails:
        ok, msg = send_email_with_attachments(
            to_emails=to_emails,
            subject=subject,
            body_text=body,
            attachments=attachments,
            cc_emails=cc_emails,
            bcc_emails=bcc_emails,
        )
        if ok:
            print(f" Email OK: {msg}")
        else:
            print(f" Email NOT sent: {msg}")
    else:
        print(" CELODOCS_EMAIL_TO no está configurado, no envío correo.")

    print(f"OK\n- JSON: {json_out}\n- XLSX: {out_xlsx}\n- Subject: {subject}")
    if sql_zip_path and sql_zip_path.exists():
        print(f"- SQL ZIP: {sql_zip_path}")

if __name__ == "__main__":
    main()